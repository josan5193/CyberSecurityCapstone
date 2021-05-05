#In order to work unless openpyxl is already installed on your pc
#You must type 'pip install openpyxl'
#Also make sure the database file is in same file as code
from openpyxl import *
import sha256
#used to mass enter users into database
#Warning using this function resets database names so be careful use only once
def create_voterdb():
    workbook = Workbook()
    sheet = workbook.active

    usernum = int(input('Enter number of users to load into database: '))
    for x in range(usernum):
        print('Name ',x+1)
        name = input('Enter Name: ')
        sheet["A"+str((x+1))] = name

    workbook.save(filename="Reg_VoterDB.xlsx")
    
#Used to place the presidential votes for each registered voter
def creat_voterselect():
    workbook2 =  workbook2()
    sheet2 = workbook2.active
    
    presnum = int(input('Enter presidential canidate number: '))
    for x in range(presnum):
        print('Number ', x+1)
        number = input('Enter number: ')
        sheet2["A"+str((x=1))] = number
       
    workbook2.save(filename="Reg_VoterSelect.xlsx")
    
def creat_hashfile():
    workbook3 =  workbook3()
    sheet3 = workbook3.active
    
    HASHEDSTRING = "print(HashedString.hexdigest())"
    for x in range(presnum):
        print('Number ', x+1)
        number = input('Enter number: ')
        sheet2["A"+str((x=1))] = number
       
    workbook3.save(filename="Reg_HashFile.xlsx")


#Reads the name database and returns it to main function
def read_db():
    wb = load_workbook("Reg_VoterDB.xlsx")
    #print(wb.sheetnames)
    sheet = wb['Sheet']
    #print(sheet['A1'].value)
    #print(sheet['A3'].value)
    #print(sheet['A3'].value)
    return wb

#Reads the number for each voter selection and return it to main
def read_db():
    wb2 = load_workbook2("Reg_VoterSelect.xlsx")
    #print(wb.sheetnames)
    sheet2 = wb2['Sheet 2']
    #print(sheet2['A1'].value)
    #print(sheet2['A3'].value)
    #print(sheet2['A3'].value)
    return wb2

def read_db():
    wb3 = load_workbook3("Reg_HashFile.xlsx")
    #print(wb.sheetnames)
    sheet3 = wb3['Sheet 3']
    #print(sheet2['A1'].value)
    #print(sheet2['A3'].value)
    #print(sheet2['A3'].value)
    return wb3

#Writes all contents of the modified wb to the database file
def write_db(wb):
    wb.save("Reg_VoterDB.xlsx")
   
#Writes all contents of the modified wb2 to the database file
def write_db(wb2):
    wb2.save("Reg_VoterSelect.xlsx")
    
def write_db(wb3):
    wb3.save("Reg_VoterSelect.xlsx")


def testmain():
    wb = read_db()
    sheet = wb['Sheet']
    print(sheet['A3'].value)
    sheet['A4'] = 'Test'
    write_db(wb)
    
    wb2 = read_db()
    sheet2 = wb2['sheet 2']
    print(sheet['A3'].value)
    sheet['A4'] = '2'
    write_db(wb2)
    
        wb3 = read_db()
    sheet3 = wb3['sheet 3']
    print(sheet['A3'].value)
    sheet['A4'] = '3'
    write_db(wb3)

    
    
    
    
